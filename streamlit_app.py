#!/usr/bin/env python3
"""
Cochise County Master Gardener Association -- Seed Library
Streamlit Web Application
Run with:  streamlit run streamlit_app.py
"""

import os
import io
import sys
import sqlite3
import json
import tempfile
import random
import smtplib
from email.mime.text import MIMEText
from datetime import datetime, timedelta
import streamlit as st

# -------------------------------------------------------------
# PAGE CONFIG (must be first Streamlit call)
# -------------------------------------------------------------
# Load custom icon
try:
    from PIL import Image as _PILImage
    import os as _os
    _icon_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "SaguaroFlower.png")
    if _os.path.exists(_icon_path):
        _page_icon = _PILImage.open(_icon_path)
    else:
        _page_icon = "🌵"
except Exception:
    _page_icon = "🌵"

st.set_page_config(
    page_title="CCMGA Seed Library",
    page_icon=_page_icon,
    layout="centered",
    initial_sidebar_state="collapsed",
)

# -------------------------------------------------------------
# PASSWORD PROTECTION
# -------------------------------------------------------------
def check_password():
    """
    Full auth gate:
      - Existing users: email + password login
      - New users: register -> email verify -> wait for admin approval
      - Admins: access admin panel in sidebar
    Returns True if authenticated and approved.
    """
    # Already authenticated this session
    if st.session_state.get("authenticated"):
        return True

    # Auth step router
    step = st.session_state.get("auth_step", "login")

    # ── Centered card ────────────────────────────────────────────────
    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown("""
        <div style="background:#1b5e20;padding:24px 32px;border-radius:12px;
                    text-align:center;margin-bottom:16px;">
          <h2 style="color:white;margin:0;">🌵 CCMGA Seed Library</h2>
          <p style="color:#c8e6c9;margin:4px 0 0 0;">
            Cochise County Master Gardener Association</p>
        </div>""", unsafe_allow_html=True)

        if step == "login":
            _auth_login(col)
        elif step == "register":
            _auth_register(col)
        elif step == "verify":
            _auth_verify(col)

    st.stop()
    return False


def _auth_login(col):
    st.markdown("#### Sign In")
    email    = st.text_input("Email", key="login_email")
    password = st.text_input("Password", type="password", key="login_pw")

    if st.button("Log In", width='stretch', type="primary"):
            if not email or not password:
                st.error("Please enter email and password.")
            else:
                user = user_login(email, password)
                if user is None:
                    st.error("Invalid email or password.")
                elif not user["is_verified"]:
                    st.warning("Email not verified. Check your inbox.")
                    st.session_state.auth_email = email
                    st.session_state.auth_step  = "verify"
                    st.rerun()
                elif not user["is_approved"]:
                    st.warning(
                        "🔒 Account verified but awaiting admin approval. "
                        "You will be notified when access is granted.")
                else:
                    update_last_login(email)
                    st.session_state.authenticated = True
                    st.session_state.user_email    = email
                    st.session_state.user_name     = user["full_name"]
                    st.session_state.user_role     = user["role"]
                    st.session_state.pop("current_page", None)
                    st.session_state.pop("_prev_selected", None)
                    st.rerun()
    if st.button("Register", width='stretch'):
        st.session_state.auth_step = "register"
        st.rerun()


def _auth_register(col):
    st.markdown("#### Create Account")
    name     = st.text_input("Full Name",            key="reg_name")
    email    = st.text_input("Email",                key="reg_email")
    password = st.text_input("Password",             key="reg_pw",   type="password")
    confirm  = st.text_input("Confirm Password",     key="reg_pw2",  type="password")

    if st.button("Sign Up", width='stretch', type="primary"):
            if not all([name, email, password, confirm]):
                st.error("All fields are required.")
            elif password != confirm:
                st.error("Passwords do not match.")
            elif len(password) < 8:
                st.error("Password must be at least 8 characters.")
            else:
                with st.spinner("Creating account..."):
                    result = user_register(email, name, password)
                if result == "ok":
                    st.session_state.auth_email = email.lower().strip()
                    st.session_state.auth_step  = "verify"
                    st.success("Account created! Check your email for the verification code.")
                    st.rerun()
                elif result == "duplicate":
                    st.error("That email is already registered. Try logging in.")
                elif result.startswith("email_failed"):
                    detail = result.replace("email_failed: ", "")
                    st.warning(f"Account created but email failed: {detail}\n\nCheck SMTP secrets. For Gmail use an App Password.")
                    st.session_state.auth_email = email.lower().strip()
                    st.session_state.auth_step  = "verify"
                    st.rerun()
                else:
                    st.error(f"Registration error: {result}")
    if st.button("Back to Login", width='stretch'):
        st.session_state.auth_step = "login"
        st.rerun()


def _auth_verify(col):
    email = st.session_state.get("auth_email", "")
    st.markdown("#### Verify Your Email")
    st.info(f"A 6-digit code was sent to **{email}**")
    code = st.text_input("Enter 6-digit code", key="verify_code", max_chars=6)

    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("Verify", width='stretch', type="primary"):
            if not code or len(code) != 6:
                st.error("Enter the 6-digit code.")
            else:
                result = user_verify(email, code.strip())
                if result == "ok":
                    st.success(
                        "✅ Email verified! Your account is pending "
                        "admin approval. You will receive an email when approved.")
                    st.session_state.auth_step = "login"
                    st.rerun()
                elif result == "expired":
                    st.error("Code expired. Request a new one.")
                else:
                    st.error("Invalid code. Please try again.")
                    
    with c2:
        # Fixed: Changed width="stretch" to width='stretch'
        if st.button("Resend Code", width='stretch'):
            with st.spinner("Sending..."):
                success = user_resend_code(email)
            # Fixed: Moved the success check inside the button scope
            if success:
                st.success("New code sent!")
            else:
                st.error("Failed to send. Contact admin.")
        
    with c3:
        if st.button("Back to Login", width='stretch'):
            st.session_state.auth_step = "login"
            st.rerun()


# -------------------------------------------------------------
# GLOBAL STYLES
# -------------------------------------------------------------
st.markdown("""
<style>
    /* ── Mobile-first responsive layout ── */

    /* Sidebar -- green on all devices */
    [data-testid="stSidebar"] {
        background-color: #1b5e20 !important;
    }
    [data-testid="stSidebar"] * {
        color: white !important;
    }
    [data-testid="stSidebar"] .stRadio label {
        color: white !important;
        font-weight: bold;
        font-size: 1.1rem;
        padding: 6px 0;
    }
    /* Sidebar toggle button -- make it visible and easy to tap */
    [data-testid="stSidebarCollapsedControl"] {
        background-color: #1b5e20 !important;
        border-radius: 0 8px 8px 0 !important;
    }
    [data-testid="stSidebarCollapsedControl"] svg {
        color: white !important;
        fill: white !important;
    }

    /* Page title header */
    .ccmga-title {
        background-color: #1b5e20;
        color: white;
        padding: 14px 16px;
        border-radius: 8px;
        margin-bottom: 16px;
        text-align: center;
    }
    .ccmga-title h1 {
        color: white;
        margin: 0;
        font-size: clamp(1.1rem, 4vw, 1.6rem);
        line-height: 1.3;
    }
    .ccmga-title p {
        color: #c8e6c9;
        margin: 4px 0 0 0;
        font-size: clamp(0.8rem, 3vw, 0.95rem);
    }

    /* Blue action buttons -- larger touch targets on mobile */
    .stButton > button {
        background-color: #0076DB !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: bold !important;
        padding: 10px 16px !important;
        min-height: 44px !important;
        font-size: clamp(0.85rem, 3vw, 1rem) !important;
        width: 100% !important;
    }
    .stButton > button:hover {
        background-color: #005aaa !important;
    }
    .stButton > button:active {
        background-color: #004080 !important;
    }

    /* Form inputs -- larger on mobile */
    .stTextInput input,
    .stNumberInput input,
    .stSelectbox select {
        min-height: 44px !important;
        font-size: 1rem !important;
    }
    .stTextArea textarea {
        font-size: 1rem !important;
    }

    /* Dataframe -- scrollable on mobile */
    .stDataFrame {
        border: 1px solid #ddd;
        border-radius: 6px;
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch;
    }

    /* Alert boxes */
    .stAlert { border-radius: 6px; }

    /* Main container -- full width on mobile */
    .block-container {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
        max-width: 100% !important;
    }

    /* Checkbox -- larger tap target */
    .stCheckbox > label {
        min-height: 36px !important;
        align-items: center !important;
        display: flex !important;
    }

    /* Download button */
    .stDownloadButton > button {
        background-color: #2e7d32 !important;
        color: white !important;
        border-radius: 8px !important;
        font-weight: bold !important;
        min-height: 44px !important;
        width: 100% !important;
    }

    /* Radio buttons in nav */
    .stRadio > div {
        gap: 8px !important;
    }

    /* On small screens collapse multi-column forms to single column */
    @media (max-width: 640px) {
        .row-widget.stHorizontal {
            flex-wrap: wrap !important;
        }
        .row-widget.stHorizontal > div {
            flex: 1 1 100% !important;
            min-width: 100% !important;
        }
        .ccmga-title h1 { font-size: 1.2rem !important; }
        .stButton > button { font-size: 1rem !important; }
    }
</style>
""", unsafe_allow_html=True)


# -------------------------------------------------------------
# DATABASE -- loaded once per session into st.session_state
# -------------------------------------------------------------
COLS = [
    "FileNumber", "Family", "Variety", "SeedSource", "Comments",
    "NumSeeds", "Season", "SeedSaverLevel", "HybridDoNotSave",
    "Edible", "WhereGrown", "PerennialAnnual", "GrownBy",
    "Year", "SoilTemperature", "Germination", "BackgroundInfo",
]

COL_HEADER_MAP = {
    "FileNumber":      "FileNumber",
    "Family":          "Family",
    "Variety":         "Variety",
    "SeedSource":      "Seed Source",
    "Comments":        "Comments",
    "NumSeeds":        "# of Seeds",
    "Season":          "Season",
    "SeedSaverLevel":  "Seed Saver Level",
    "HybridDoNotSave": "Hybrid-Do Not Save",
    "Edible":          "Edible",
    "WhereGrown":      "Where Grown",
    "PerennialAnnual": "Perennial/Annual",
    "GrownBy":         "Grown By",
    "Year":            "Year",
    "SoilTemperature": "Soil Temperature",
    "Germination":     "Germination",
    "BackgroundInfo":  "Background Info",
}
SHEET_HEADERS = list(COL_HEADER_MAP.values())

CREATE_SQL = """
    CREATE TABLE IF NOT EXISTS seeds (
        "FileNumber"      INTEGER PRIMARY KEY,
        "Family"          TEXT,
        "Variety"         TEXT,
        "SeedSource"      TEXT,
        "Comments"        TEXT,
        "NumSeeds"        TEXT,
        "Season"          TEXT,
        "SeedSaverLevel"  TEXT,
        "HybridDoNotSave" TEXT,
        "Edible"          TEXT,
        "WhereGrown"      TEXT,
        "PerennialAnnual" TEXT,
        "GrownBy"         TEXT,
        "Year"            TEXT,
        "SoilTemperature" TEXT,
        "Germination"     TEXT,
        "BackgroundInfo"  TEXT
    )
"""

# -------------------------------------------------------------
# POSTGRESQL DATABASE LAYER
# -------------------------------------------------------------

import streamlit as st
import psycopg2

def get_pg_conn():
    import psycopg2
    import psycopg2.extras

    url = st.secrets["DATABASE_URL"]

    try:
        conn = psycopg2.connect(
            url,
            connect_timeout=10,
            cursor_factory=psycopg2.extras.RealDictCursor,
        )
        conn.autocommit = False
        return conn

    except Exception as e:
        st.exception(e)
        raise
    
CREATE_USERS_SQL = """
    CREATE TABLE IF NOT EXISTS app_users (
        id             SERIAL PRIMARY KEY,
        email          TEXT UNIQUE NOT NULL,
        full_name      TEXT,
        password_hash  TEXT NOT NULL,
        is_verified    BOOLEAN DEFAULT FALSE,
        is_approved    BOOLEAN DEFAULT FALSE,
        verify_code    TEXT,
        verify_expires TIMESTAMP,
        created_at     TIMESTAMP DEFAULT NOW(),
        last_login     TIMESTAMP,
        role           TEXT DEFAULT 'user'
    )
"""


def _ensure_table():
    """Create the seeds and app_users tables if they don't exist."""
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute(CREATE_SQL)
    cur.execute(CREATE_USERS_SQL)
    conn.commit()
    cur.close()
    conn.close()


def _seed_table_populated() -> bool:
    """Return True if the seeds table has at least one row."""
    try:
        conn = get_pg_conn()
        cur  = conn.cursor()
        cur.execute('SELECT 1 FROM seeds LIMIT 1')
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row is not None
    except Exception:
        return False


def _load_from_xlsx_to_pg():
    """
    One-time seed load: read the local xlsx and INSERT into PostgreSQL.
    Only runs if the table is empty.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path  = None
    for d in [script_dir, os.path.dirname(script_dir), os.getcwd()]:
        cand = os.path.join(d, "_SEED_LIBRARY_PARSED.xlsx")
        if os.path.exists(cand):
            xlsx_path = cand
            break

    if not xlsx_path:
        st.session_state["db_status"] = "warning"
        st.session_state["db_msg"] = (
            "⚠️ PostgreSQL table is empty and no xlsx found to seed it. "
            "Upload _SEED_LIBRARY_PARSED.xlsx alongside the app for initial load.")
        return

    try:
        import openpyxl
        wb       = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws       = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not all_rows:
            raise ValueError("Spreadsheet is empty.")

        headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]

        def ci(name):
            try:    return headers.index(name)
            except: return -1

        idx = {k: ci(v) for k, v in COL_HEADER_MAP.items()}

        def gv(rv, key):
            i = idx.get(key, -1)
            if i < 0 or i >= len(rv): return ""
            v = rv[i]
            return str(v).strip() if v is not None else ""

        conn = get_pg_conn()
        cur  = conn.cursor()
        inserted = 0
        for rv in all_rows[1:]:
            fn_raw = gv(rv, "FileNumber")
            if not fn_raw or fn_raw == "None": continue
            try:    fn = int(float(fn_raw))
            except: continue
            yr = gv(rv, "Year")
            try:    yr = str(int(float(yr))) if yr and yr != "None" else ""
            except: pass
            cur.execute("""
                INSERT INTO seeds
                ("FileNumber","Family","Variety","SeedSource","Comments",
                 "NumSeeds","Season","SeedSaverLevel","HybridDoNotSave",
                 "Edible","WhereGrown","PerennialAnnual","GrownBy","Year",
                 "SoilTemperature","Germination","BackgroundInfo")
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT ("FileNumber") DO NOTHING
            """, (fn,
                  gv(rv,"Family"),      gv(rv,"Variety"),
                  gv(rv,"SeedSource"),  gv(rv,"Comments"),
                  gv(rv,"NumSeeds"),    gv(rv,"Season"),
                  gv(rv,"SeedSaverLevel"), gv(rv,"HybridDoNotSave"),
                  gv(rv,"Edible"),      gv(rv,"WhereGrown"),
                  gv(rv,"PerennialAnnual"), gv(rv,"GrownBy"), yr,
                  gv(rv,"SoilTemperature"), gv(rv,"Germination"),
                  gv(rv,"BackgroundInfo")))
            inserted += 1
        conn.commit()
        cur.close()
        conn.close()
        st.session_state["db_status"] = "ok"
        st.session_state["db_msg"] = f"✅ Loaded {inserted} seeds into PostgreSQL."
    except Exception as e:
        import traceback; traceback.print_exc()
        st.session_state["db_status"] = "error"
        st.session_state["db_msg"] = f"[x] Initial load error: {e}"


def init_db():
    """Called once per session -- ensures table exists and has data."""
    if st.session_state.get("db_ready"):
        return
    try:
        _ensure_table()
        st.success("Database tables initialized successfully")
    except Exception as e:
        st.exception(e)
        st.stop()


# -- CRUD -----------------------------------------------------
def sf(row: dict, key: str) -> str:
    """Safely get a string field from a PostgreSQL row -- never returns None."""
    v = row.get(key) if isinstance(row, dict) else None
    return str(v).strip() if v is not None else ""


def db_search(term: str = "") -> list[dict]:
    conn = get_pg_conn()
    cur  = conn.cursor()
    if term:
        t = f"%{term.lower()}%"
        cur.execute("""
            SELECT * FROM seeds
            WHERE LOWER(CAST("FileNumber" AS TEXT)) LIKE %s
               OR LOWER("Family")  LIKE %s
               OR LOWER("Variety") LIKE %s
            ORDER BY "FileNumber"
        """, (t, t, t))
    else:
        cur.execute('SELECT * FROM seeds ORDER BY "FileNumber"')
    # Coerce None ? "" but keep FileNumber as int
    def clean_row(r):
        d = dict(r)
        return {k: (int(v) if k == "FileNumber" and v is not None
                    else str(v).strip() if v is not None else "")
                for k, v in d.items()}
    rows = [clean_row(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def db_add(data: dict):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        INSERT INTO seeds
        ("FileNumber","Family","Variety","SeedSource","Comments",
         "NumSeeds","Season","SeedSaverLevel","HybridDoNotSave",
         "Edible","WhereGrown","PerennialAnnual","GrownBy","Year",
         "SoilTemperature","Germination","BackgroundInfo")
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, tuple(data.get(c, "") for c in COLS))
    conn.commit()
    cur.close()
    conn.close()


def db_update(fn: int, data: dict):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        UPDATE seeds SET
            "Family"=%s, "Variety"=%s, "SeedSource"=%s, "Comments"=%s,
            "NumSeeds"=%s, "Season"=%s, "SeedSaverLevel"=%s,
            "HybridDoNotSave"=%s, "Edible"=%s, "WhereGrown"=%s,
            "PerennialAnnual"=%s, "GrownBy"=%s, "Year"=%s,
            "SoilTemperature"=%s, "Germination"=%s, "BackgroundInfo"=%s
        WHERE "FileNumber"=%s
    """, (
        data.get("Family",""),        data.get("Variety",""),
        data.get("SeedSource",""),    data.get("Comments",""),
        data.get("NumSeeds",""),      data.get("Season",""),
        data.get("SeedSaverLevel",""),data.get("HybridDoNotSave",""),
        data.get("Edible",""),        data.get("WhereGrown",""),
        data.get("PerennialAnnual",""),data.get("GrownBy",""),
        data.get("Year",""),          data.get("SoilTemperature",""),
        data.get("Germination",""),   data.get("BackgroundInfo",""),
        fn,
    ))
    conn.commit()
    cur.close()
    conn.close()


def db_delete(file_numbers: list[int]):
    conn = get_pg_conn()
    cur  = conn.cursor()
    for fn in file_numbers:
        cur.execute('DELETE FROM seeds WHERE "FileNumber"=%s', (fn,))
    conn.commit()
    cur.close()
    conn.close()


def db_next_fn() -> int:
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute('SELECT MAX("FileNumber") FROM seeds')
    row = cur.fetchone()
    cur.close()
    conn.close()
    val = row["max"] if row else None
    return (val + 1) if val else 1001


def db_count() -> int:
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM seeds")
    row = cur.fetchone()
    cur.close()
    conn.close()
    return row["count"] if row else 0


def db_over_limit() -> list[dict]:
    """Return seeds where Comments or BackgroundInfo exceeds 300 chars."""
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT "FileNumber", "Family", "Variety",
               LENGTH("Comments")       AS clen,
               LENGTH("BackgroundInfo") AS blen
        FROM seeds
        WHERE LENGTH("Comments") > 300
           OR LENGTH("BackgroundInfo") > 300
        ORDER BY "Family", "Variety"
    """)
    rows = []
    for r in cur.fetchall():
        d = dict(r)
        rows.append({
            "FileNumber": d.get("FileNumber"),
            "Family":     str(d.get("Family") or ""),
            "Variety":    str(d.get("Variety") or ""),
            "clen":       int(d.get("clen") or 0),
            "blen":       int(d.get("blen") or 0),
        })
    cur.close()
    conn.close()
    return rows


def save_to_xlsx():
    """
    Build an in-memory xlsx from PostgreSQL data and store in session
    for download. No file writing needed -- PG is the source of truth.
    """
    try:
        import openpyxl
        rows = db_search("")
        wb   = openpyxl.Workbook()
        ws   = wb.active
        ws.title = "Seeds"
        ws.append(SHEET_HEADERS)
        for row in rows:
            ws.append([row.get(k, "") for k in COL_HEADER_MAP.keys()])
        buf = io.BytesIO()
        wb.save(buf)
        st.session_state["xlsx_download_bytes"] = buf.getvalue()
    except Exception as e:
        print(f"xlsx build error: {e}")


# =============================================================
# USER AUTHENTICATION FUNCTIONS
# =============================================================

def _hash_password(plain: str) -> str:
    import bcrypt
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()


def _check_password_hash(plain: str, hashed: str) -> bool:
    import bcrypt
    try:
        return bcrypt.checkpw(plain.encode(), hashed.encode())
    except Exception:
        return False


def _send_verification_email(to_email: str, code: str) -> tuple:
    """Send 6-digit code via SMTP. Returns (success, error_message)."""
    required = ["SMTP_SERVER", "SMTP_PORT", "EMAIL_FROM", "EMAIL_PASSWORD"]
    missing = [k for k in required if k not in st.secrets]
    if missing:
        return False, f"Missing secrets: {', '.join(missing)}"
    try:
        msg = MIMEText(
            f"Your 6-digit verification code for the CCMGA Seed Library is:\n\n"
            f"    {code}\n\nThis code expires in 15 minutes."
        )
        msg["Subject"] = "Verify Your CCMGA Seed Library Account"
        msg["From"]    = st.secrets["EMAIL_FROM"]
        msg["To"]      = to_email
        with smtplib.SMTP(st.secrets["SMTP_SERVER"],
                          int(st.secrets["SMTP_PORT"])) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(st.secrets["EMAIL_FROM"], st.secrets["EMAIL_PASSWORD"])
            server.send_message(msg)
        return True, ""
    except Exception as e:
        return False, str(e)


def user_login(email: str, password: str) -> dict | None:
    """Return user dict if credentials valid, else None."""
    conn = get_pg_conn()
    cur = conn.cursor()

    cur.execute(
        "SELECT * FROM app_users WHERE email = %s",
        (email.lower().strip(),)
    )

    row = cur.fetchone()

    cur.close()
    conn.close()

    if row is None:
        return None

    # Convert the returned row to a normal dictionary
    user = dict(row)

    if not _check_password_hash(password, user["password_hash"]):
        return None

    return user

def user_register(email: str, full_name: str, password: str) -> str:
    """
    Insert new unverified user. Returns 'ok', 'duplicate', or 'error'.
    """
    code    = str(random.randint(100000, 999999))
    expires = (datetime.now() + timedelta(minutes=15)).isoformat()
    try:
        conn = get_pg_conn()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO app_users
                (email, full_name, password_hash,
                 is_verified, is_approved, verify_code, verify_expires, role)
            VALUES (%s, %s, %s, FALSE, FALSE, %s, %s, 'user')
        """, (email.lower().strip(), full_name,
              _hash_password(password), code, expires))
        conn.commit()
        _send_admin_notification(
    full_name,
    email
)
        cur.close()
        conn.close()
        # Send email
        ok, err = _send_verification_email(email, code)
        if ok:
            return "ok"
        return f"email_failed: {err}"
    except Exception as e:
        if "unique" in str(e).lower() or "duplicate" in str(e).lower():
            return "duplicate"
        return f"error: {e}"


def user_verify(email: str, code: str) -> str:
    """
    Verify the 6-digit code. Returns 'ok', 'expired', or 'invalid'.
    """
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("SELECT * FROM app_users WHERE email = %s",
                (email.lower().strip(),))
    row = cur.fetchone()
    if not row:
        cur.close(); conn.close()
        return "invalid"
    user = dict(row)
    if user["verify_code"] != code:
        cur.close(); conn.close()
        return "invalid"
    # Use timezone-aware comparison to handle PostgreSQL timestamptz
    from datetime import timezone
    expires_raw = user["verify_expires"]
    if expires_raw is not None:
        if hasattr(expires_raw, "tzinfo") and expires_raw.tzinfo is not None:
            now = datetime.now(timezone.utc)
        else:
            now = datetime.now()
        if expires_raw < now:
            cur.close(); conn.close()
            return "expired"
    cur.execute(
        """
        UPDATE app_users
        SET
            is_verified = TRUE,
            verify_code = NULL,
            verify_expires = NULL
        WHERE email = %s
        """,
        (email.lower().strip(),)
    )
    conn.commit()
    cur.close()
    conn.close()
    return "ok"
    

def user_resend_code(email: str) -> bool:
    """Generate a new code and resend verification email."""
    code    = str(random.randint(100000, 999999))
    expires = (datetime.now() + timedelta(minutes=15)).isoformat()
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        UPDATE app_users
        SET verify_code = %s, verify_expires = %s
        WHERE email = %s
    """, (code, expires, email.lower().strip()))
    conn.commit()
    cur.close()
    conn.close()
    success, _ = _send_verification_email(
        email,
        code
)
    return success

def admin_get_pending() -> list[dict]:
    """Return all verified-but-not-approved users."""
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT id, email, full_name, created_at
        FROM app_users
        WHERE is_verified = TRUE AND is_approved = FALSE
        ORDER BY created_at
    """)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def admin_get_all_users() -> list[dict]:
    """Return all users for admin management."""
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("""
        SELECT id, email, full_name, is_verified, is_approved,
               role, created_at, last_login
        FROM app_users ORDER BY created_at DESC
    """)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def admin_approve(user_id: int):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute(
        "UPDATE app_users SET is_approved = TRUE WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()


def admin_revoke(user_id: int):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute(
        "UPDATE app_users SET is_approved = FALSE WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()


def admin_delete_user(user_id: int):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute("DELETE FROM app_users WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()


def update_last_login(email: str):
    conn = get_pg_conn()
    cur  = conn.cursor()
    cur.execute(
        "UPDATE app_users SET last_login = NOW() WHERE email = %s",
        (email.lower().strip(),))
    conn.commit()
    cur.close()
    conn.close()

def _send_admin_notification(name,email):

    admin_email = st.secrets["ADMIN_EMAIL"]

    msg = MIMEText(
        f"""
New account awaiting approval

Name:

{name}

Email:

{email}

Login to admin panel.
"""
    )

    msg["Subject"]="New User Registration"

    msg["From"]=st.secrets["EMAIL_FROM"]

    msg["To"]=admin_email

    try:

        with smtplib.SMTP(
            st.secrets["SMTP_SERVER"],
            int(st.secrets["SMTP_PORT"])
        ) as srv:

            srv.starttls()

            srv.login(
                st.secrets["EMAIL_FROM"],
                st.secrets["EMAIL_PASSWORD"]
            )

            srv.send_message(msg)

    except:

        pass
def generate_labels_pdf(label_data: list,
                        include_background: bool = False) -> bytes | None:
    """Returns PDF bytes or None on error.
    If include_background=True, appends a separate background-info label
    for each seed that has BackgroundInfo text.
    Hybrid warning is printed on every label where HybridDoNotSave is set.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units    import inch
        from reportlab.lib          import colors
        from reportlab.platypus     import Paragraph, Frame
        from reportlab.lib.styles   import ParagraphStyle
        from reportlab.lib.enums    import TA_CENTER, TA_LEFT
        from reportlab.pdfgen       import canvas as rl_canvas
    except ImportError:
        st.error("ReportLab not installed. Run:  pip install reportlab")
        return None

    labels = []      # (row, is_bg_label)
    for row, qty in label_data:
        for _ in range(qty):
            labels.append((row, False))
        # Add same qty of background labels as seed labels (only if bg info exists)
        if include_background and (row.get("BackgroundInfo") or "").strip():
            for _ in range(qty):
                labels.append((row, True))
    if not labels:
        return None

    # Avery 94207 exact dimensions
    PAGE_W, PAGE_H  = letter              # 8.5 x 11 inches
    MARGIN_TOP      = 0.50 * inch
    MARGIN_LEFT     = 0.25 * inch
    MARGIN_RIGHT    = 0.25 * inch
    LABEL_W         = 4.00 * inch         # hard-coded, not calculated
    LABEL_H         = 2.00 * inch
    GUTTER          = 0.25 * inch
    LEFT_X          = 0.25 * inch         # left edge of col 0
    RIGHT_X         = 4.375 * inch        # 0.25 + 4.00 + 0.125
    COLS, ROWS      = 2, 5
    PER_PAGE        = COLS * ROWS

    PAD_L, PAD_R, PAD_T, PAD_B = 4, 4, 3, 3
    TITLE_H         = 24
    LEFT_FRAC       = 2 / 3

    BORDER  = colors.HexColor("#000000")
    DIVIDER = colors.HexColor("#888888")
    GREEN   = colors.HexColor("#225522")

    title_sty = ParagraphStyle("ttl", fontSize=10, fontName="Helvetica-Bold",
        textColor=GREEN, alignment=TA_CENTER, leading=12, spaceAfter=0)
    fam_sty = ParagraphStyle("fam", fontSize=10, fontName="Helvetica-Bold",
        textColor=colors.red, alignment=TA_LEFT, leading=12, spaceAfter=1)
    var_sty = ParagraphStyle("var", fontSize=10, fontName="Helvetica-Oblique",
        textColor=colors.black, alignment=TA_LEFT, leading=12, spaceAfter=2)
    cmt_sty = ParagraphStyle("cmt", fontSize=9, fontName="Helvetica",
        textColor=colors.black, alignment=TA_LEFT, leading=11, spaceAfter=0)
    rgt_sty = ParagraphStyle("rgt", fontSize=9, fontName="Helvetica",
        textColor=colors.black, alignment=TA_CENTER, leading=11, spaceAfter=1)
    rit_sty = ParagraphStyle("rit", fontSize=9, fontName="Helvetica-Oblique",
        textColor=colors.black, alignment=TA_CENTER, leading=11, spaceAfter=1)
    svr_sty = ParagraphStyle("svr", fontSize=7, fontName="Helvetica-Bold",
        textColor=colors.black, alignment=TA_CENTER, leading=9,
        spaceAfter=1, wordWrap="LTR")
    grm_sty = ParagraphStyle("grm", fontSize=8, fontName="Helvetica",
        textColor=colors.black, alignment=TA_CENTER, leading=10, spaceAfter=0)

    buf = io.BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=letter)

    page_idx = 0
    while page_idx * PER_PAGE < len(labels):
        page_labels = labels[page_idx * PER_PAGE : (page_idx + 1) * PER_PAGE]

        for slot, (row, is_bg) in enumerate(page_labels):
            col_num = slot % COLS
            row_num = slot // COLS
            lx = LEFT_X if col_num == 0 else RIGHT_X
            ly = PAGE_H - MARGIN_TOP - (row_num + 1) * LABEL_H

            # Extract fields (needed for both label types)
            family   = (row.get("Family")          or "").strip()
            variety  = (row.get("Variety")         or "").strip()
            season   = (row.get("Season")          or "").strip()
            numseeds = (row.get("NumSeeds")        or "").strip()
            edible   = (row.get("Edible")          or "").strip()
            year_val = (row.get("Year")            or "").strip()
            comment  = " ".join((row.get("Comments") or "").split())[:300]
            saver    = (row.get("SeedSaverLevel")  or "").strip()
            germ     = (row.get("Germination")     or "").strip()
            soil_t   = (row.get("SoilTemperature") or "").strip()
            hybrid   = (row.get("HybridDoNotSave") or "").strip()
            bg_info  = " ".join((row.get("BackgroundInfo") or "").split())[:300]

            # No borders on any label
            # Common layout measurements
            tx     = lx + PAD_L
            tw     = LABEL_W - PAD_L - PAD_R
            body_y = ly + PAD_B
            body_x = lx + PAD_L
            body_w = LABEL_W - PAD_L - PAD_R

            if is_bg:
                # -- Background Info label -- clean, no dividers, full width --
                # Generous padding for readability
                BG_PAD = 10
                full_w = LABEL_W - BG_PAD * 2
                full_h = LABEL_H - BG_PAD * 2

                # Title: "Family -- Background Info"
                bg_title_sty = ParagraphStyle("bgtitle",
                    fontSize=10, fontName="Helvetica-Bold",
                    textColor=colors.black, alignment=TA_LEFT,
                    leading=13, spaceAfter=4)
                bg_body_sty = ParagraphStyle("bgbody",
                    fontSize=9, fontName="Helvetica",
                    textColor=colors.black, alignment=TA_LEFT,
                    leading=12, spaceAfter=0)

                all_bg = [
                    Paragraph(f"{family}", bg_title_sty),
                    Paragraph(f"{variety} -- Background Information", bg_title_sty),
                    Paragraph(bg_info, bg_body_sty),
                ]
                bg_frame = Frame(lx + BG_PAD, ly + BG_PAD, full_w, full_h,
                      leftPadding=0, rightPadding=0,
                      topPadding=0, bottomPadding=0,
                      showBoundary=0)
                c.saveState()
                _clip = c.beginPath()
                _clip.rect(lx + BG_PAD, ly + BG_PAD, full_w, full_h)
                c.clipPath(_clip, stroke=0, fill=0)
                bg_frame.addFromList(all_bg, c)
                c.restoreState()

            else:
                # -- Standard seed label ----------------------------
                TITLE_H_USE = TITLE_H
                ty     = ly + LABEL_H - PAD_T - TITLE_H_USE
                body_h = LABEL_H - PAD_T - TITLE_H_USE - 2 - PAD_B
                left_w = body_w * LEFT_FRAC
                right_w = body_w * (1 - LEFT_FRAC)
                vdiv_x = body_x + left_w + 2

                # Header
                title_frame = Frame(tx, ty, tw, TITLE_H_USE, leftPadding=0, rightPadding=0,
                      topPadding=0, bottomPadding=0, showBoundary=0)
                c.saveState()
                _clip = c.beginPath()
                _clip.rect(tx, ty, tw, TITLE_H_USE)
                c.clipPath(_clip, stroke=0, fill=0)
                title_frame.addFromList(
                    [Paragraph("Cochise County Master Gardener Association"
                               "<br/>Seed Library", title_sty)], c)
                c.restoreState()

                # Thin horizontal divider under header
                div_y = ly + LABEL_H - PAD_T - TITLE_H_USE - 1
                c.setStrokeColor(DIVIDER)
                c.setLineWidth(0.5)
                c.line(lx + 2, div_y, lx + LABEL_W - 2, div_y)

                # Vertical divider between left and right cells
                c.line(vdiv_x, ly + PAD_B + 2, vdiv_x, div_y - 2)

                # Left cell
                left_items = [Paragraph(family, fam_sty)]
                if variety: left_items.append(Paragraph(variety, var_sty))
                if hybrid:  left_items.append(Paragraph(
                    f"* HYBRID -- DO NOT SAVE SEEDS *",
                    ParagraphStyle("hyb", fontSize=7,
                    fontName="Helvetica-Bold",
                    textColor=colors.HexColor("#b71c1c"),
                    alignment=TA_LEFT, leading=9)))
                if comment: left_items.append(Paragraph(comment, cmt_sty))
                left_frame = Frame(body_x, body_y, left_w - 4, body_h,
                      leftPadding=0, rightPadding=0,
                      topPadding=0, bottomPadding=0, showBoundary=0)
                c.saveState()
                _clip = c.beginPath()
                _clip.rect(body_x, body_y, left_w - 4, body_h)
                c.clipPath(_clip, stroke=0, fill=0)
                left_frame.addFromList(left_items, c)
                c.restoreState()

                # Right cell
                right_items = []
                if year_val: right_items.append(Paragraph(year_val, rgt_sty))
                if edible:   right_items.append(Paragraph(edible.upper(), rgt_sty))
                if season:   right_items.append(Paragraph(season, rit_sty))
                if numseeds: right_items.append(Paragraph(f"{numseeds} Seeds", rgt_sty))
                if saver:    right_items.append(Paragraph(saver, svr_sty))
                # Show germ and soil temp as separate clean lines
                if germ:
                    right_items.append(Paragraph(f"Germ: {germ}", grm_sty))
                if soil_t:
                    right_items.append(Paragraph(f"Soil: {soil_t}", grm_sty))
                right_frame = Frame(vdiv_x + 3, body_y, right_w - 4, body_h,
                      leftPadding=0, rightPadding=0,
                      topPadding=0, bottomPadding=0, showBoundary=0)
                c.saveState()
                _clip = c.beginPath()
                _clip.rect(vdiv_x + 3, body_y, right_w - 4, body_h)
                c.clipPath(_clip, stroke=0, fill=0)
                right_frame.addFromList(right_items, c)
                c.restoreState()

        c.showPage()
        page_idx += 1

    c.save()
    return buf.getvalue()


# -------------------------------------------------------------
# SHARED UI HELPERS
# -------------------------------------------------------------
def show_download_bar():
    """Show download button if updated xlsx bytes are available."""
    xlsx_bytes = st.session_state.get("xlsx_download_bytes")
    if xlsx_bytes:
        st.download_button(
            label="  Download Updated _SEED_LIBRARY_PARSED.xlsx",
            data=xlsx_bytes,
            file_name="_SEED_LIBRARY_PARSED.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download and replace your local xlsx file to keep changes permanent.",
            width='stretch',
        )


def page_header(title: str, subtitle: str = ""):
    st.markdown(f"""
    <div class="ccmga-title">
        <h1>🌵 {title}</h1>
        {"<p>" + subtitle + "</p>" if subtitle else ""}
    </div>
    """, unsafe_allow_html=True)


FIELD_LABELS = {
    "FileNumber":      "File Number",
    "Family":          "Family",
    "Variety":         "Variety",
    "SeedSource":      "Seed Source",
    "Comments":        "Comments",
    "NumSeeds":        "# of Seeds",
    "Season":          "Season",
    "SeedSaverLevel":  "Seed Saver Level",
    "HybridDoNotSave": "Hybrid-Do Not Save",
    "Edible":          "Edible",
    "WhereGrown":      "Where Grown",
    "PerennialAnnual": "Perennial/Annual",
    "GrownBy":         "Grown By",
    "Year":            "Year",
    "SoilTemperature": "Soil Temperature",
    "Germination":     "Germination",
    "BackgroundInfo":  "Background Info",
}

SEASON_OPTS   = ["", "Cool Season", "Warm Season", "Hot Season",
                 "All Season", "Decorative"]
SAVER_OPTS    = ["", "Beginner Seed Saver", "Easy Seed Saver",
                 "Intermediate Seed Saver", "Advanced Seed Saver"]
PERAN_OPTS    = ["", "Annual", "Perennial", "Perennial/Annual"]


# -------------------------------------------------------------
# PAGE: HOME
# -------------------------------------------------------------
def page_home():
    page_header(
        "Cochise County Master Gardener Association",
        "Seed Library Database",
    )
    init_db()  # ensure DB is loaded
    status = st.session_state.get("db_status", "")
    msg    = st.session_state.get("db_msg", "")
    if status == "ok":
        st.success(msg)
    elif status == "warning":
        st.warning(msg)
    elif status == "error":
        st.error(msg)

    count = db_count()
    st.markdown(f"### 🌱 {count:,} seeds in the library")
    show_download_bar()
    st.markdown("---")

    st.markdown("""
**Use the navigation menu on the left sidebar to move between pages.**

| | |
|---|---|
| **Built for** | Cochise County Master Gardener Association |
| **Labels** | For use with Avery 94207 Labels (2" x 4", 10 per sheet) |
| **Version** | 3.0 7.12.2026 |
    """)

    # -- Seeds with comments or background info over 300 chars -------
    over_limit = db_over_limit()

    if over_limit:
        st.markdown("---")
        st.warning(f"⚠️ **{len(over_limit)} seed(s) have text exceeding the "
                   "300-character label limit.** Only the first 300 characters "
                   "will print. Edit these records to shorten the text.")
        rows_display = []
        for r in over_limit:
            issues = []
            clen = int(r.get("clen") or 0)
            blen = int(r.get("blen") or 0)
            if clen > 300:
                issues.append(f"Comments: {clen} chars ({clen-300} over)")
            if blen > 300:
                issues.append(f"Background Info: {blen} chars ({blen-300} over)")
            rows_display.append({
                "File #":  r.get("FileNumber", ""),
                "Family":  sf(r, "Family"),
                "Variety": sf(r, "Variety"),
                "Issue":   "  |  ".join(issues),
            })
        import pandas as pd
        st.dataframe(
            pd.DataFrame(rows_display),
            width='stretch',
            hide_index=True,
        )


# -------------------------------------------------------------
# PAGE: BROWSE
# -------------------------------------------------------------
def page_browse():
    page_header("Browse Seeds", "Search, sort, view details, and edit records")

    # Search bar
    term = st.text_input("Search", placeholder="Family, variety, or file number...",
                         key="browse_search")
    c_s1, c_s2 = st.columns(2)
    with c_s1:
        search_clicked = st.button("Search", width='stretch')
    with c_s2:
        show_all = st.button("Show All", width='stretch')
    if show_all:
        st.session_state.browse_term = ""
    elif search_clicked:
        st.session_state.browse_term = term

    active_term = st.session_state.get("browse_term", "")
    rows = db_search(active_term)

    # Deduplicate by Family+Variety -- keep first occurrence, count duplicates
    seen      = {}
    unique    = []
    count_map = {}
    for r in rows:
        key = (sf(r,"Family").lower(), sf(r,"Variety").lower())
        count_map[key] = count_map.get(key, 0) + 1
    for r in rows:
        key = (sf(r,"Family").lower(), sf(r,"Variety").lower())
        if key not in seen:
            seen[key] = True
            unique.append(r)

    st.caption(f"{len(unique)} unique varieties  ({len(rows)} total records)")

    if not unique:
        st.info("No seeds found.")
        return

    # Display table with import pandas trick-free approach
    import pandas as pd
    display_cols = ["FileNumber", "Family", "Variety", "Count",
                    "Season", "NumSeeds", "SeedSaverLevel",
                    "PerennialAnnual", "GrownBy", "Year"]

    table_data = []
    for r in unique:
        key = (sf(r,"Family").lower(), sf(r,"Variety").lower())
        row_d = {c: r.get(c, "") for c in display_cols if c != "Count"}
        row_d["Count"] = count_map[key]
        table_data.append(row_d)

    df = pd.DataFrame(table_data)[
        ["FileNumber", "Family", "Variety", "Count",
         "Season", "NumSeeds", "SeedSaverLevel", "PerennialAnnual",
         "GrownBy", "Year"]
    ]
    df.columns = ["File #", "Family", "Variety", "Count",
                  "Season", "# Seeds", "Saver Level",
                  "Perennial/Annual", "Grown By", "Year"]

    st.dataframe(df, width='stretch', hide_index=True,
                 height=320)

    st.markdown("---")

    # -- Detail / Edit panel -------------------------------------
    st.markdown("#### View or Edit a Record")
    fn_options = {f"#{r['FileNumber']}  {r['Family']} -- {r['Variety']}": r["FileNumber"]
                  for r in unique}
    chosen_label = st.selectbox("Select seed", list(fn_options.keys()),
                                key="browse_select")
    chosen_fn = fn_options[chosen_label]
    chosen_row = next((r for r in rows if r["FileNumber"] == chosen_fn), None)

    if chosen_row:
        # Nav-style action buttons
        if "browse_action" not in st.session_state:
            st.session_state.browse_action = "View"
        ba1, ba2, ba3 = st.columns(3)
        with ba1:
            if st.button("View",
                         width='stretch',
                         type="primary" if st.session_state.browse_action == "View" else "secondary"):
                st.session_state.browse_action = "View"
                st.rerun()
        with ba2:
            if st.button("Edit",
                         width='stretch',
                         type="primary" if st.session_state.browse_action == "Edit" else "secondary"):
                st.session_state.browse_action = "Edit"
                st.rerun()
        with ba3:
            if st.button("Duplicate",
                         width='stretch',
                         type="primary" if st.session_state.browse_action == "Duplicate as New Record" else "secondary"):
                st.session_state.browse_action = "Duplicate as New Record"
                st.rerun()
        st.markdown("---")
        action = st.session_state.browse_action
        if action == "View":
            _browse_detail(chosen_row)
        elif action == "Edit":
            _browse_edit_form(chosen_row, is_duplicate=False)
        elif action == "Duplicate as New Record":
            _browse_duplicate_form(chosen_row)


def _browse_detail(row: dict):
    """Read-only detail view -- Comments and Background Info side by side."""
    # Top: seed identity + quick facts
    id_fields    = ["Family", "Variety", "SeedSource", "GrownBy", "WhereGrown"]
    right_fields = ["FileNumber", "Year", "Season", "NumSeeds", "Edible",
                    "PerennialAnnual", "SeedSaverLevel",
                    "SoilTemperature", "Germination", "HybridDoNotSave"]

    col1 = st.container()
    col2 = st.container()
    with col1:
        for f in id_fields:
            st.markdown(f"**{FIELD_LABELS[f]}:** {row.get(f,'') or '--'}")
    with col2:
        for f in right_fields:
            st.markdown(f"**{FIELD_LABELS[f]}:** {row.get(f,'') or '--'}")

    # Comments and Background Info side by side (or stacked on mobile)
    comments = str(row.get("Comments") or "").strip()
    bg       = str(row.get("BackgroundInfo") or "").strip()

    if comments or bg:
        st.markdown("---")
        if comments and bg:
            cc, bc = st.columns(2)
        else:
            cc = bc = st.container()

        if comments:
            with cc:
                st.markdown("**Comments**")
                st.info(comments)
                if len(comments) > 300:
                    st.warning(
                        f"⚠️ {len(comments)} chars "
                        f"({len(comments)-300} over 300-char label limit)")
        if bg:
            with bc:
                st.markdown("**Background Information**")
                st.info(bg)
                if len(bg) > 300:
                    st.warning(
                        f"⚠️ {len(bg)} chars "
                        f"({len(bg)-300} over 300-char label limit)")


def _browse_edit_form(row: dict, is_duplicate: bool = False):
    """Editable form for the selected row."""
    fn = row["FileNumber"]
    st.markdown(f"**Editing File #{fn}**")
    with st.form(key=f"edit_form_{fn}_{is_duplicate}"):
        c1 = st.container()
        c2 = st.container()
        with c1:
            family  = st.text_input("Family",       value=row.get("Family",""))
            variety = st.text_input("Variety",       value=row.get("Variety",""))
            source  = st.text_input("Seed Source",   value=row.get("SeedSource",""))
            comments= st.text_area("Comments",       value=row.get("Comments",""), height=100,
                                   help="Max 300 chars for label printing", max_chars=300)
            grown_by= st.text_input("Grown By",      value=row.get("GrownBy",""))
            where   = st.text_input("Where Grown",   value=row.get("WhereGrown",""))
        with c2:
            year    = st.text_input("Year",          value=row.get("Year",""))
            numseeds= st.text_input("# of Seeds",    value=row.get("NumSeeds",""))
            edible  = st.text_input("Edible",        value=row.get("Edible",""))
            season_v= row.get("Season","")
            season  = st.selectbox("Season", SEASON_OPTS,
                                   index=SEASON_OPTS.index(season_v)
                                   if season_v in SEASON_OPTS else 0)
            saver_v = row.get("SeedSaverLevel","")
            saver   = st.selectbox("Seed Saver Level", SAVER_OPTS,
                                   index=SAVER_OPTS.index(saver_v)
                                   if saver_v in SAVER_OPTS else 0)
            peran_v = row.get("PerennialAnnual","")
            peran   = st.selectbox("Perennial/Annual", PERAN_OPTS,
                                   index=PERAN_OPTS.index(peran_v)
                                   if peran_v in PERAN_OPTS else 0)
            hybrid  = st.text_input("Hybrid-Do Not Save", value=row.get("HybridDoNotSave",""))
            soil_t  = st.text_input("Soil Temperature",   value=row.get("SoilTemperature",""))
            germ    = st.text_input("Germination",        value=row.get("Germination",""))
        bg_info = st.text_area("Background Info",
                               value=row.get("BackgroundInfo",""), height=80,
                               help="Max 300 chars for label printing", max_chars=300)

        if st.form_submit_button("Save Changes", width='stretch'):
            db_update(fn, {
                "Family": family, "Variety": variety,
                "SeedSource": source, "Comments": comments,
                "NumSeeds": numseeds, "Season": season,
                "SeedSaverLevel": saver, "HybridDoNotSave": hybrid,
                "Edible": edible, "WhereGrown": where,
                "PerennialAnnual": peran, "GrownBy": grown_by,
                "Year": year, "SoilTemperature": soil_t,
                "Germination": germ, "BackgroundInfo": bg_info,
            })
            st.success(f"Seed #{fn} updated successfully.")
            st.rerun()


def _browse_duplicate_form(source_row: dict):
    """Duplicate a seed record with a new file number."""
    next_fn = db_next_fn()
    st.info(f"Duplicating **#{source_row['FileNumber']} {source_row['Family']} -- "
            f"{source_row['Variety']}** as new record **#{next_fn}**. "
            "Edit any fields then save.")
    with st.form(key=f"dup_form_{source_row['FileNumber']}"):
        c1 = st.container()
        c2 = st.container()
        with c1:
            fn      = st.number_input("File Number *", value=next_fn,
                                      min_value=1, step=1)
            family  = st.text_input("Family",     value=source_row.get("Family",""))
            variety = st.text_input("Variety",    value=source_row.get("Variety",""))
            source  = st.text_input("Seed Source",value=source_row.get("SeedSource",""))
            comments= st.text_area("Comments",    value=source_row.get("Comments",""), height=80)
            grown_by= st.text_input("Grown By",   value=source_row.get("GrownBy",""))
            where   = st.text_input("Where Grown",value=source_row.get("WhereGrown",""))
        with c2:
            year    = st.text_input("Year",       value=source_row.get("Year",""))
            numseeds= st.text_input("# of Seeds", value=source_row.get("NumSeeds",""))
            edible  = st.text_input("Edible",     value=source_row.get("Edible",""))
            season_v = source_row.get("Season","")
            season  = st.selectbox("Season", SEASON_OPTS,
                                   index=SEASON_OPTS.index(season_v)
                                   if season_v in SEASON_OPTS else 0)
            saver_v = source_row.get("SeedSaverLevel","")
            saver   = st.selectbox("Seed Saver Level", SAVER_OPTS,
                                   index=SAVER_OPTS.index(saver_v)
                                   if saver_v in SAVER_OPTS else 0)
            peran_v = source_row.get("PerennialAnnual","")
            peran   = st.selectbox("Perennial/Annual", PERAN_OPTS,
                                   index=PERAN_OPTS.index(peran_v)
                                   if peran_v in PERAN_OPTS else 0)
            hybrid  = st.text_input("Hybrid-Do Not Save",
                                    value=source_row.get("HybridDoNotSave",""))
            soil_t  = st.text_input("Soil Temperature",
                                    value=source_row.get("SoilTemperature",""))
            germ    = st.text_input("Germination",
                                    value=source_row.get("Germination",""))
        bg_info = st.text_area("Background Info",
                               value=source_row.get("BackgroundInfo",""), height=80)

        if st.form_submit_button("Save as New Record", width='stretch'):
            fn = int(fn)
            _conn = get_pg_conn()
            _cur  = _conn.cursor()
            _cur.execute('SELECT 1 FROM seeds WHERE "FileNumber"=%s', (fn,))
            existing = _cur.fetchone()
            _cur.close()
            _conn.close()
            if existing:
                st.error(f"File #{fn} already exists. Choose a different number.")
            else:
                db_add({
                    "FileNumber": fn, "Family": family, "Variety": variety,
                    "SeedSource": source, "Comments": comments,
                    "NumSeeds": numseeds, "Season": season,
                    "SeedSaverLevel": saver, "HybridDoNotSave": hybrid,
                    "Edible": edible, "WhereGrown": where,
                    "PerennialAnnual": peran, "GrownBy": grown_by,
                    "Year": year, "SoilTemperature": soil_t,
                    "Germination": germ, "BackgroundInfo": bg_info,
                })
                st.success(f"✅ New seed #{fn} saved as a duplicate of "
                           f"#{source_row['FileNumber']}.")
                show_download_bar()
                st.rerun()


# -------------------------------------------------------------
# PAGE: ADD SEEDS
# -------------------------------------------------------------
def page_add():
    page_header("Add Seeds", "Enter a new seed record into the library")

    next_fn = db_next_fn()
    st.info(f"Next available File Number: **{next_fn}**")

    with st.form("add_form", clear_on_submit=True):
        c1 = st.container()
        c2 = st.container()
        with c1:
            fn      = st.number_input("File Number *", value=next_fn,
                                      min_value=1, step=1)
            family  = st.text_input("Family")
            variety = st.text_input("Variety")
            source  = st.text_input("Seed Source")
            comments= st.text_area("Comments", height=100,
                                   help="Maximum 300 characters for label printing",
                                   max_chars=300)
            grown_by= st.text_input("Grown By")
            where   = st.text_input("Where Grown")
        with c2:
            year    = st.text_input("Year")
            numseeds= st.text_input("# of Seeds")
            edible  = st.text_input("Edible (type 'Edible' if applicable)")
            season  = st.selectbox("Season",           SEASON_OPTS)
            saver   = st.selectbox("Seed Saver Level", SAVER_OPTS)
            peran   = st.selectbox("Perennial/Annual", PERAN_OPTS)
            hybrid  = st.text_input("Hybrid-Do Not Save")
            soil_t  = st.text_input("Soil Temperature")
            germ    = st.text_input("Germination")
        bg_info = st.text_area("Background Info", height=80,
                               help="Maximum 300 characters for label printing",
                               max_chars=300)

        submitted = st.form_submit_button("Save Seed", width='stretch')

    if submitted:
        fn = int(fn)
        if not family:
            st.error("Family is required.")
        else:
            # Check duplicate
            _conn = get_pg_conn()
            _cur  = _conn.cursor()
            _cur.execute('SELECT 1 FROM seeds WHERE "FileNumber"=%s', (fn,))
            existing = _cur.fetchone()
            _cur.close()
            _conn.close()
            if existing:
                st.error(f"File #{fn} already exists. Choose a different number.")
            else:
                db_add({
                    "FileNumber": fn, "Family": family, "Variety": variety,
                    "SeedSource": source, "Comments": comments,
                    "NumSeeds": numseeds, "Season": season,
                    "SeedSaverLevel": saver, "HybridDoNotSave": hybrid,
                    "Edible": edible, "WhereGrown": where,
                    "PerennialAnnual": peran, "GrownBy": grown_by,
                    "Year": year, "SoilTemperature": soil_t,
                    "Germination": germ, "BackgroundInfo": bg_info,
                })
                st.success(f"✅ Seed #{fn} -- {family} added successfully!")
                show_download_bar()


# -------------------------------------------------------------
# PAGE: REMOVE SEEDS
# -------------------------------------------------------------
def page_remove():
    page_header("Remove Seeds", "Search for and permanently delete seed records")

    term = st.text_input("Search", placeholder="Family, variety, or file number...",
                         key="remove_search")
    c_r1, c_r2 = st.columns(2)
    with c_r1:
        if st.button("Search", width='stretch'):
            st.session_state.remove_term = term
    with c_r2:
        if st.button("Show All", width='stretch'):
            st.session_state.remove_term = ""

    active_term = st.session_state.get("remove_term", "")
    rows = db_search(active_term)

    if not rows:
        st.info("No seeds found.")
        return

    import pandas as pd
    df = pd.DataFrame([{
        "Select": False,
        "File #": r["FileNumber"],
        "Family": sf(r,"Family"),
        "Variety": sf(r,"Variety"),
        "Season": sf(r,"Season"),
        "Year": sf(r,"Year"),
    } for r in rows])

    st.caption(f"{len(rows)} records found. Check boxes to select for deletion.")

    edited = st.data_editor(
        df,
        width='stretch',
        hide_index=True,
        height=380,
        column_config={
            "Select": st.column_config.CheckboxColumn(
                "Select", help="Check to mark for deletion", default=False),
        },
        disabled=["File #", "Family", "Variety", "Season", "Year"],
        key="remove_editor",
    )

    selected_fns = [
        int(edited.iloc[i]["File #"])
        for i in range(len(edited))
        if edited.iloc[i]["Select"]
    ]
    n = len(selected_fns)

    if n > 0:
        st.warning(f"**{n} record(s) selected for deletion.**")
        confirm = st.checkbox(
            f"I confirm I want to permanently delete {n} seed record(s)",
            key="remove_confirm")
        if confirm:
            if st.button(f"Delete {n} Selected Record(s)",
                         type="primary", width='stretch'):
                db_delete(selected_fns)
                st.success(f"✅ {n} record(s) deleted.")
                show_download_bar()
                st.session_state.remove_term = active_term
                st.rerun()
    else:
        st.info("Check the Select box on one or more rows to delete them.")


# -------------------------------------------------------------
# PAGE: PRINT LABELS
# -------------------------------------------------------------
def page_labels():
    page_header("Print Seed Labels",
                "Avery 94207 -- 2 inch x 4 inch labels, 10 per sheet (2 cols x 5 rows)")

    # Session state init
    if "label_term"       not in st.session_state: st.session_state.label_term       = ""
    if "label_qtys"       not in st.session_state: st.session_state.label_qtys       = {}
    if "label_include_bg" not in st.session_state: st.session_state.label_include_bg = False
    if "label_pdf_bytes"  not in st.session_state: st.session_state.label_pdf_bytes  = None

    # Fetch rows once
    rows       = db_search(st.session_state.label_term)
    row_lookup = {int(r["FileNumber"]): r for r in rows}

    # Build label_data from current selections
    label_data   = []
    total_labels = 0
    for fn, qty in st.session_state.label_qtys.items():
        if qty > 0 and fn in row_lookup:
            label_data.append((row_lookup[fn], qty))
            total_labels += qty
    n_seeds = len(label_data)

    # ═══════════════════════════════════════════════════════════════
    # TOP PANEL: summary + generate + download  (always visible)
    # ═══════════════════════════════════════════════════════════════
    st.markdown(
        f"<div style='background:#e3f2fd;padding:10px 14px;border-radius:8px;"
        f"margin-bottom:10px;font-size:1.05rem;'>"
        f"<b>{n_seeds}</b> seed type(s) selected &nbsp;|&nbsp; "
        f"<b>{total_labels}</b> total labels</div>",
        unsafe_allow_html=True)

    btn1, btn2, btn3 = st.columns([3, 3, 2])
    with btn1:
        gen_clicked = st.button(
            "Generate PDF",
            type="primary",
            width='stretch',
            disabled=(n_seeds == 0),
            key="gen_pdf_btn")
    with btn2:
        if st.session_state.label_pdf_bytes:
            st.download_button(
                label="Download seed_labels.pdf",
                data=st.session_state.label_pdf_bytes,
                file_name="seed_labels.pdf",
                mime="application/pdf",
                width='stretch',
                key="dl_pdf_btn")
        else:
            st.button("Download PDF", disabled=True,
                      width='stretch', key="dl_pdf_btn_off")
    with btn3:
        bg_active = st.session_state.label_include_bg
        if st.button(
                "BG: ON" if bg_active else "BG: OFF",
                width='stretch',
                type="primary" if bg_active else "secondary",
                key="bg_toggle"):
            st.session_state.label_include_bg = not bg_active
            st.session_state.label_pdf_bytes  = None
            st.rerun()

    if gen_clicked:
        with st.spinner("Generating PDF..."):
            pdf_bytes = generate_labels_pdf(
                label_data,
                include_background=st.session_state.label_include_bg)
        if pdf_bytes:
            st.session_state.label_pdf_bytes = pdf_bytes
            pages = -(-total_labels // 10)
            st.success(
                f"PDF ready -- {total_labels} labels across {pages} page(s). "
                "Print at Actual Size (100%). Click Download above.")
            st.rerun()

    if st.session_state.label_pdf_bytes and not gen_clicked:
        pages = -(-total_labels // 10)
        st.caption(f"Last PDF: {total_labels} labels / {pages} page(s). "
                   "Re-generate if selections changed.")

    st.divider()

    # ═══════════════════════════════════════════════════════════════
    # SEARCH + BULK ACTIONS
    # ═══════════════════════════════════════════════════════════════
    term = st.text_input("Search seeds",
                         placeholder="Family, variety, or file number...",
                         key="label_search")
    sa, sb, sc, sd = st.columns(4)
    with sa:
        if st.button("Search", width='stretch', key="lbl_srch"):
            st.session_state.label_term      = term
            st.session_state.label_pdf_bytes = None
            st.rerun()
    with sb:
        if st.button("Load All", width='stretch', key="lbl_all"):
            st.session_state.label_term      = ""
            st.session_state.label_pdf_bytes = None
            st.rerun()
    with sc:
        if st.button("Set All to 1", width='stretch', key="lbl_set1"):
            for r in rows:
                st.session_state.label_qtys[int(r["FileNumber"])] = 1
            st.session_state.label_pdf_bytes = None
            st.rerun()
    with sd:
        if st.button("Clear All", width='stretch', key="lbl_clr"):
            st.session_state.label_qtys      = {}
            st.session_state.label_pdf_bytes = None
            st.rerun()

    if not rows:
        st.info("No seeds found. Click Load All or try a different search.")
        return

    st.caption(f"{len(rows)} seed(s) in list. Qty >= 1 = included in PDF.")

    # ═══════════════════════════════════════════════════════════════
    # SEED SELECTION LIST
    # ═══════════════════════════════════════════════════════════════
    hdr = st.columns([1, 7, 2])
    hdr[0].markdown("**Sel**")
    hdr[1].markdown("**Family / Variety**")
    hdr[2].markdown("**Qty**")
    st.divider()

    for r in rows:
        fn      = int(r["FileNumber"])
        cur_qty = st.session_state.label_qtys.get(fn, 0)
        cols    = st.columns([1, 7, 2])
        checked = cols[0].checkbox(
            "", value=cur_qty > 0,
            key=f"lbl_chk_{fn}",
            label_visibility="collapsed")
        cols[1].markdown(
            f"**{sf(r,'Family')}**  \n{sf(r,'Variety')}  "
            f"<span style='color:#888;font-size:0.8em'>{sf(r,'Season')}</span>",
            unsafe_allow_html=True)
        new_qty = cols[2].number_input(
            "", min_value=0, max_value=99,
            value=max(cur_qty, 1 if checked else 0),
            step=1, key=f"lbl_qty_{fn}",
            label_visibility="collapsed")
        if checked and new_qty == 0:
            new_qty = 1
        if not checked:
            new_qty = 0
        st.session_state.label_qtys[fn] = new_qty


def page_admin():
    """Admin panel -- manage user approvals and accounts."""
    if st.session_state.get("user_role") != "admin":
        st.error("Access denied.")
        return

    page_header("Admin Panel", "Manage user accounts and approvals")

    tab1, tab2 = st.tabs(["Pending Approvals", "All Users"])

    with tab1:
        pending = admin_get_pending()
        if not pending:
            st.success("No accounts waiting for approval.")
        else:
            st.warning(f"{len(pending)} account(s) awaiting approval.")
            for u in pending:
                c1, c2, c3 = st.columns([3, 1, 1])
                c1.markdown(
                    f"**{u['full_name']}**  \n{u['email']}  \n"
                    f"*Registered: {str(u.get('created_at',''))[:10]}*")
                if c2.button("Approve", key=f"apr_{u['id']}",
                             width='stretch', type="primary"):
                    admin_approve(u["id"])
                    st.session_state["current_page"] = "admin"
                    # Notify user by email
                    try:
                        msg = MIMEText(
                            f"Hello {u['full_name']},\n\n"
                            "Your CCMGA Seed Library account has been approved!\n"
                            "You can now log in at the app URL.\n\n"
                            "Cochise County Master Gardener Association")
                        msg["Subject"] = "CCMGA Seed Library -- Account Approved"
                        msg["From"]    = st.secrets["EMAIL_FROM"]
                        msg["To"]      = u["email"]
                        with smtplib.SMTP(st.secrets["SMTP_SERVER"],
                                          int(st.secrets["SMTP_PORT"])) as srv:
                            srv.starttls()
                            srv.login(st.secrets["EMAIL_FROM"],
                                      st.secrets["EMAIL_PASSWORD"])
                            srv.send_message(msg)
                    except Exception:
                        pass
                    st.rerun()
                if c3.button("Deny", key=f"deny_{u['id']}",
                             width='stretch'):
                    admin_delete_user(u["id"])
                    st.session_state["current_page"] = "admin"
                    st.rerun()
                st.divider()

    with tab2:
        all_users = admin_get_all_users()
        import pandas as pd
        df = pd.DataFrame([{
            "Name":       u.get("full_name",""),
            "Email":      u.get("email",""),
            "Role":       u.get("role","user"),
            "Verified":   "Yes" if u.get("is_verified") else "No",
            "Approved":   "Yes" if u.get("is_approved") else "No",
            "Registered": str(u.get("created_at",""))[:10],
            "Last Login":  str(u.get("last_login",""))[:10] if u.get("last_login") else "Never",
        } for u in all_users])
        st.dataframe(df, width='stretch', hide_index=True)

        st.markdown("---")
        st.markdown("#### Manage User")
        emails = [u["email"] for u in all_users]
        sel_email = st.selectbox("Select user", emails, key="admin_sel")
        sel_user  = next((u for u in all_users if u["email"] == sel_email), None)
        if sel_user:
            c1, c2, c3 = st.columns(3)
            if sel_user["is_approved"]:
                if c1.button("Revoke Access", width='stretch'):
                    admin_revoke(sel_user["id"])
                    st.rerun()
            else:
                if c1.button("Approve", width='stretch',
                             type="primary"):
                    admin_approve(sel_user["id"])
                    st.rerun()
            new_role = c2.selectbox(
                "Role", ["user","admin"],
                index=0 if sel_user["role"] == "user" else 1,
                key="role_sel")
            if c2.button("Set Role", width='stretch'):
                conn = get_pg_conn()
                cur  = conn.cursor()
                cur.execute(
                    "UPDATE app_users SET role = %s WHERE id = %s",
                    (new_role, sel_user["id"]))
                conn.commit(); cur.close(); conn.close()
                st.rerun()
            if c3.button("Delete User", width='stretch'):
                if sel_user["role"] != "admin":
                    admin_delete_user(sel_user["id"])
                    st.rerun()
                else:
                    st.error("Cannot delete an admin account.")


def sidebar_nav():
    PAGES = ["Home", "Browse Seeds", "Add Seeds", "Remove Seeds", "Print Labels"]

    target = st.session_state.pop("nav_target", None)
    if target and target in PAGES:
        st.session_state["_nav_index"] = PAGES.index(target)
    current_idx = st.session_state.get("_nav_index", 0)

    with st.sidebar:
        st.markdown("## 🌵 CCMGA\n### Seed Library")
        # Show logged-in user
        uname = st.session_state.get("user_name", "")
        urole = st.session_state.get("user_role", "user")
        if uname:
            st.markdown(f"**{uname}**  "
                        f"{'(Admin)' if urole == 'admin' else ''}")
        st.markdown("---")

        page = st.radio(
            "Navigate",
            PAGES,
            index=current_idx,
            key="nav_radio",
            label_visibility="collapsed",
        )
        st.session_state["_nav_index"] = PAGES.index(page)
        st.markdown("---")

        # Admin panel link
        if urole == "admin":
            is_on_admin = st.session_state.get("current_page") == "admin"
            if st.button("Admin Panel" + (" (active)" if is_on_admin else ""),
                         width='stretch'):
                st.session_state["current_page"] = "admin"
                st.rerun()

        st.markdown(
            "<small>Cochise County Master Gardener Association<br/>"
            "v3.0 7.12.2026 Alan Borhauer</small>",
            unsafe_allow_html=True,
        )
        st.markdown("---")
        if st.button("🔒 Log Out", width='stretch'):
            for k in ["authenticated","user_email","user_name",
                      "user_role","auth_step","auth_email"]:
                st.session_state.pop(k, None)
            st.rerun()
    return page


# -------------------------------------------------------------
# MAIN
# -------------------------------------------------------------
def main():
    # Gate -- must authenticate before anything else renders
    check_password()

    selected = sidebar_nav()

    # "current_page" persists until user clicks a sidebar radio item
    # Clicking a radio item clears current_page so we go back to normal nav
    # We detect a radio change by storing the previous selection
    prev = st.session_state.get("_prev_selected", selected)
    if selected != prev:
        # User clicked a different radio item -- leave admin mode
        st.session_state.pop("current_page", None)
    st.session_state["_prev_selected"] = selected

    current_page = st.session_state.get("current_page", "")
    if current_page == "admin":
        page_admin()
        return

    if selected == "Home":
        page_home()
    elif selected == "Browse Seeds":
        page_browse()
    elif selected == "Add Seeds":
        page_add()
    elif selected == "Remove Seeds":
        page_remove()
    elif selected == "Print Labels":
        page_labels()


if __name__ == "__main__":
    main()
